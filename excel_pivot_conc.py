import sys
import platform
import flask
import re
import logging
import os

from flask import Response
from io import BytesIO
from flask import Flask, send_file

from database import db
from flask import render_template
from flask.blueprints import Blueprint
from flask import current_app
from flask import Flask, session
from flask import request, flash

dbinstance = Blueprint('dbinstance', __name__,
                 template_folder='templates',
                 static_folder='static')
                 
from model import Base
import bcrypt

import os

import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.orm import relationship, Session
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.inspection import inspect
from sqlalchemy.dialects import postgresql
#from sqlalchemy import db.db.db.mapped_column, Table, Column, Bigdb.Integer, db.Integer, db.db.Text, db.Date, db.Boolean, db.db.db.String, ForeignKey
from sqlalchemy import desc, text

import csv
import io
#from io import StringIO
from flask import make_response

import csv
from io import StringIO
#import StringIO       # allows you to store response object in memory instead of on disk
from flask import Flask, make_response # Necessary imports, should be obvious

from sqlalchemy import text

from openpyxl import Workbook
from flask import Response
from sqlalchemy import inspect

from sqlalchemy import column
import datetime
#from flask_http_response import success, result, error

#from app.models import DataModelTbl #inherits from DB.Model
#from db import DataModelTbl #inherits from DB.Model

from flask import Response
from openpyxl import Workbook
#from openpyxl.writer.excel import save_virtual_workbook
import flask_excel as excel


@dbinstance.route('/download_sheet')
def download():
    create_excel_sheet(data)
    output = excel.make_response()
    output.headers["Content-Disposition"] = "attachment; filename=sheet.xlsx"
    output.headers["Content-type"] = "application/vnd.openxmlformats-\
officedocument.spreadsheetml.sheet"
    return output  # Send the response to the user
#<a href="{{ url_for('app.download') }}">Click here for Sheet</a>

@dbinstance.route('/excel_pivot_conc')
def excel_pivot_conc():
    with db.engine.begin() as conn:
        ResultSet = conn.execute(text("select colpivot('_test_pivoted',"
"        'SELECT libelle_commune, par_idsuf, u_nom_raison_sociale, no_interne"
" FROM t_parceldem"
" INNER JOIN t_demande"
" ON t_parceldem.par_nointerne = t_demande.no_interne"
" LEFT JOIN t_commune"
" ON CASE WHEN SUBSTRING(t_parceldem.par_idsuf, 6, 3) LIKE ''000''"
" THEN SUBSTRING(t_parceldem.par_idsuf, 1, 5)"
" ELSE CONCAT (SUBSTRING(t_parceldem.par_idsuf, 1, 2),"
" SUBSTRING(t_parceldem.par_idsuf, 6, 3))"
" END  = t_commune.code_insee_commune "
" INNER JOIN t_usager"
" ON t_demande.no_pacage_demandeur = t_usager.u_pacage"
" WHERE par_idsuf like ''22193000ZH0016%''"
" GROUP BY t_demande.date_complet, t_parceldem.par_idsuf,"
" t_demande.no_interne, t_usager.u_pacage, t_usager.u_nom_raison_sociale,"
" t_commune.libelle_commune, t_parceldem.par_surface"
" ORDER BY par_idsuf asc, no_interne desc,"
" libelle_commune asc LIMIT 100', "
" array['libelle_commune', 'par_idsuf'], array['no_interne', 'u_nom_raison_sociale'],"
" '#.par_idsuf', null);"
" select * from _test_pivoted order by libelle_commune, par_idsuf;"))

        rows = ResultSet.fetchall()
        #return excel.make_response_from_records(rows,"xlsx",file_name="sheet.xlsx")
        conn.commit()
        x = datetime.datetime.now()
        #response = Response(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        #Response['Content-Disposition'] = 'attachment; filename=export.xlsx'
        #response = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        #response = Response(content_type=open_xml)
        #response.headers['Content-Disposition'] = 'attachment; filename=data.xlsx'    
        #response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        #response['Content-Disposition'] = 'attachment; filename=export.xlsx'
        #open_xml = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        #response = Response(content_type=open_xml)
        #response.headers['Content-Disposition'] = 'attachment; filename=data.xlsx'
        #return response  # Send the response to the user
        #<a href="{{ url_for('app.download') }}">Click here for Sheet</a>
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'REPORT'    
        columns = [
    'libelle',
    'parcelles',
    'superficie',
    'nomproprietaireoumandataire',
    'adressenomproprietaireoumandataire',
    'demandeur',
    'adrdem',
    'cedant',
    'no_interne',
    'date_complet',
        ]
        row_num = 1
        for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                #cell.font = header_font
                #column_letter = get_column_letter(col_num)
                #column_dimensions = worksheet.column_dimensions[column_letter]
        for tris in rows:
                row_num += 1
                row = [
            tris[0],  # dv_no
            tris[1],  # account_no
            tris[2],  # id_no
            tris[3],  # first_name
            tris[4],  # middle_name
            tris[5],  # last_name
            tris[6],  # date_travel
            tris[7],  # division
            tris[8],  # division
            tris[9],  # division
                ]       
        for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                #column_letter = get_column_letter(col_num)
                #column_dimensions = worksheet.column_dimensions[column_letter]
                #column_dimensions.width = 19

        #workbook.save(response)
        #query_sets = session.query(Products).all()
        query_sets = rows
        column_names = None
        #column_names = ['title', 'type', 'components', 'region', 'producer', 'blurb', 'desc1', 'desc2', 'pairing','pretax','posttax']
        response = excel.make_response_from_query_sets(query_sets, column_names, "xls", file_name="export_data") , render_template('')
        #response = excel.make_response_from_records(rows,"xlsx",file_name="sheet.xlsx")
        return response


